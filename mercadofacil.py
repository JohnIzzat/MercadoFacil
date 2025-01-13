import openpyxl
from openpyxl import Workbook

# Nome do arquivo XLS
ARQUIVO_XLS = "base_de_dados.xlsx"


def criar_arquivo_xls():
    """Cria um arquivo XLS inicial, se não existir."""
    try:
        wb = openpyxl.load_workbook(ARQUIVO_XLS)
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.title = "Produtos"
        ws.append(["Produto", "Valor"])
        wb.save(ARQUIVO_XLS)
        print(f"Arquivo {ARQUIVO_XLS} criado com sucesso!")


def registrar_produto(produto, valor):
    """Registra um produto e valor no arquivo XLS."""
    wb = openpyxl.load_workbook(ARQUIVO_XLS)
    ws = wb["Produtos"]
    ws.append([produto, valor])
    wb.save(ARQUIVO_XLS)
    print(f"Produto '{produto}' com valor {valor} registrado com sucesso!")


def buscar_menor_valor(produto):
    """Busca o menor valor registrado para um produto."""
    wb = openpyxl.load_workbook(ARQUIVO_XLS)
    ws = wb["Produtos"]
    # Aqui já usamos diretamente row[1], pois `values_only=True` retorna os valores das células.
    valores = [
        row[1] for row in ws.iter_rows(min_row=2, values_only=True)
        if row[0] and row[0].lower() == produto.lower()
    ]
    if valores:
        menor_valor = min(valores)
        print(f"O menor valor registrado para '{produto}' é: {menor_valor}")
    else:
        print(f"Produto '{produto}' não encontrado.")


def menu():
    """Menu principal do programa."""
    criar_arquivo_xls()
    while True:
        print("\n1. Registrar produto")
        print("2. Buscar menor valor de um produto")
        print("3. Sair")
        opcao = input("Escolha uma opção: ")

        if opcao == "1":
            produto = input("Informe o nome do produto: ")
            valor = float(input("Informe o valor do produto: "))
            registrar_produto(produto, valor)
        elif opcao == "2":
            produto = input("Informe o nome do produto para buscar: ")
            buscar_menor_valor(produto)
        elif opcao == "3":
            print("Saindo do programa. Até mais!")
            break
        else:
            print("Opção inválida! Tente novamente.")


# Executa o programa
menu()
