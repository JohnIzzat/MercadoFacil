import openpyxl
from openpyxl import Workbook
import customtkinter as ctk

# variavel Janela representa a Interface gráfica

# Criando a Janela principal
janela = ctk.CTk()
janela.title("Mercado Facil")
janela.geometry("400x300")

# aparencia de acordo com a configuração do pc
ctk.set_appearance_mode("System")
# Tema de cor: pode ser blue, green, ou dark-blue
ctk.set_default_color_theme("dark-blue")


# Nome do arquivo XLS
ARQUIVO_XLS = "base_de_dados.xlsx"


def criar_arquivo_xls():
    """Cria um arquivo XLS inicial, se não existir."""
    try:
        wb = openpyxl.load_workbook(ARQUIVO_XLS)
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.title = "Produtos"
        ws.append(["Produto", "Valor"])
        wb.save(ARQUIVO_XLS)
        print(f"Arquivo {ARQUIVO_XLS} criado com sucesso!")


def registrar_produto(produto, valor):
    """Registra um produto e valor no arquivo XLS."""
    wb = openpyxl.load_workbook(ARQUIVO_XLS)
    ws = wb["Produtos"]
    ws.append([produto, valor])
    wb.save(ARQUIVO_XLS)
    print(f"Produto '{produto}' com valor {valor} registrado com sucesso!")

    def registrar_produto_gui():
        janela.withdraw
        registrar_produtoo = ctk.CTkToplevel()
        registrar_produtoo.title("Cadastrar Produtos")
        registrar_produtoo.geometry("400x300")

        def homepage():
            registrar_produtoo.destroy()
            janela.deiconify
        


# Função para buscar o menor valor de um produto e Interface

def buscar_menor_valor(produto):
    """Busca o menor valor registrado para um produto."""
    try:
        wb = openpyxl.load_workbook(ARQUIVO_XLS)
        ws = wb["Produtos"]
        valores = [
            row[1] for row in ws.iter_rows(min_row=2, values_only=True)
            if row[0] and row[0].lower() == produto.lower()
        ]
        if valores:
            menor_valor = min(valores)
            return f"O menor valor registrado para '{produto}' é: R${menor_valor:.2f}"
        else:
            return f"Produto '{produto}' não encontrado."
    except FileNotFoundError:
        return "Arquivo XLS não encontrado."
    except Exception as e:
        return f"Erro: {str(e)}"

# Funções da Interface Consultar Produto

def consultar_produto_gui():
    janela.withdraw()  # Oculta a janela principal
    consultar_produto = ctk.CTkToplevel()  # Cria a nova janela
    consultar_produto.title("Consultar Produto")
    consultar_produto.geometry("400x300")

    # Função para voltar para a tela principal
    def home():
        consultar_produto.destroy()  # Fecha a janela atual
        janela.deiconify()  # Reexibe a janela principal

    def realizar_busca():
        produto = entrada_consultarProduto.get()
        resultado = buscar_menor_valor(produto)
        label_consultarProduto.configure(text=resultado)  # Atualiza o texto da label

    # Widgets na nova janela
    label_consultarProduto = ctk.CTkLabel(consultar_produto, text="Digite o nome do produto")
    label_consultarProduto.pack(pady=20)

    entrada_consultarProduto = ctk.CTkEntry(consultar_produto, width=300, placeholder_text="Nome do Produto")
    entrada_consultarProduto.pack(pady=10)

    botao_buscarProduto = ctk.CTkButton(consultar_produto, text="Buscar", command=realizar_busca)
    botao_buscarProduto.pack(pady=20)

    label_resultadoConsultaProduto = ctk.CTkLabel(consultar_produto, text="", wraplength=300)
    label_resultadoConsultaProduto.pack(pady=10)

    botao_voltar = ctk.CTkButton(
        consultar_produto, text="Inicio", command=home
    )
    botao_voltar.pack(pady=20)



# TELA PRINCIPAL

# Botão para abrir a janela de Consulta de Produto 
botao_consultarProduto = ctk.CTkButton(
    janela, text="Consultar Produto", command=consultar_produto_gui
)
botao_consultarProduto.pack(padx=20, pady=20)

# Janela para abrir a Janela de Cadastro de Produto
botao_cadastrarProduto = ctk.CTkButton(janela, text="Cadastrar Produto", command=registrar_produto)
botao_cadastrarProduto.pack(padx=20, pady=20)


# Iniciar o loop principal da interface
janela.mainloop()
